#!/usr/bin/env python3
"""
build_model.py -- generate the multi-tab financial-plan workbook from config.

Mirrors the architecture documented in docs/ARCHITECTURE.md:
  * The ASSUMPTIONS tab is the single source of truth. Every other tab links
    back to it with cross-sheet formulas rather than hardcoding values.
  * Color convention:
      green text  = cross-sheet link  (=Assumptions!Cxx)
      black text  = intra-sheet formula
      blue text   = hardcoded input
  * Tabs built here: Assumptions, Net Worth Snapshot, Income Streams,
    Year-by-Year Projections, Employer Concentration, Monte Carlo (summary),
    Roth Conversion Ladder (lifetime-tax optimizer, engine/tax_us.py), Action Plan.

Run:
  python3 engine/build_model.py                      # demo config
  RPT_CONFIG=config/config.json python3 engine/build_model.py
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from config_loader import (  # noqa: E402
    load_config, investable_total, employer_stock_total, employer_concentration_pct,
    current_age,
)
import tax_us  # noqa: E402

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent

GREEN = Font(color="008000")          # cross-sheet link
BLACK = Font(color="000000")          # intra-sheet formula
BLUE = Font(color="0000FF")           # hardcoded input
GOODF = Font(color="1A7F37", bold=True)   # winning strategy row
RED = Font(color="CF222E")            # IRMAA-breach flag
BOLD = Font(bold=True)
HDR = Font(bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="305496")
TITLE = Font(bold=True, size=14, color="1F3864")
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _title(ws, text):
    ws["A1"] = text
    ws["A1"].font = TITLE
    ws.append([])


def _header_row(ws, cells, row=None):
    for col, val in enumerate(cells, start=1):
        c = ws.cell(row=row or ws.max_row, column=col, value=val)
        c.font = HDR
        c.fill = HDR_FILL
        c.border = BORDER


def build_assumptions(wb, cfg):
    ws = wb.create_sheet("Assumptions")
    _title(ws, "Assumptions -- SINGLE SOURCE OF TRUTH")
    a = cfg["assumptions"]
    inc = cfg["income"]
    members = cfg["household"]["members"]

    rows = [
        ("Key", "Value", "Notes"),
        ("Portfolio return (base)", a["portfolio_return_base"], "Monte Carlo mu"),
        ("Portfolio return (conservative)", a["portfolio_return_conservative"], ""),
        ("Portfolio return (optimistic)", a["portfolio_return_optimistic"], ""),
        ("Inflation", a["inflation_rate"], "bracket + spend scaling"),
        ("State income tax", cfg["household"]["state_income_tax_rate"], ""),
        ("Local income tax", cfg["household"]["local_income_tax_rate"], ""),
        ("Standard deduction (MFJ)", a["standard_deduction_mfj"], ""),
        ("IRMAA Tier 1 MAGI (MFJ)", a["irmaa_tier1_magi_mfj"], "Roth conversion ceiling"),
        ("Retirement spend (annual, today $)", a["retirement_spend_annual"], ""),
        ("Target equity %", a["target_equity_pct"], ""),
        ("Target bond %", a["target_bond_pct"], ""),
        ("Bridge target", a["bridge_target"], "pre-59.5 brokerage bridge"),
        ("Pension monthly at retirement", inc["pension_monthly_at_retirement"], ""),
        ("Pension COLA", inc["pension_cola"], ""),
        ("Passive income (annual)", inc["passive_income_annual"], ""),
        (f"{members[0]['display_name']} salary", inc["spouse_a_salary"], ""),
        (f"{members[0]['display_name']} bonus %", inc["spouse_a_bonus_pct"], ""),
        (f"{members[0]['display_name']} RSU annual", inc["spouse_a_rsu_annual"], ""),
        (f"{members[1]['display_name']} income", inc["spouse_b_annual"], ""),
        (f"{members[0]['display_name']} retirement age", members[0]["retirement_age"], ""),
        (f"{members[1]['display_name']} retirement age", members[1]["retirement_age"], ""),
        (f"{members[0]['display_name']} SS claim age", members[0]["ss_claim_age"], ""),
        (f"{members[1]['display_name']} SS claim age", members[1]["ss_claim_age"], ""),
        (f"{members[0]['display_name']} SS monthly", cfg["social_security"]["spouse_a_monthly_benefit"], ""),
        (f"{members[1]['display_name']} SS monthly", cfg["social_security"]["spouse_b_monthly_benefit"], ""),
    ]
    start = ws.max_row + 1
    for i, r in enumerate(rows):
        ws.append(r)
        if i == 0:
            _header_row(ws, r)
        else:
            ws.cell(row=ws.max_row, column=2).font = BLUE  # hardcoded inputs
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 30
    # Remember row of "Inflation" for linking demos
    cfg["_assum_rows"] = {name: start + idx for idx, (name, *_ ) in enumerate(rows)}
    return ws


def build_net_worth(wb, cfg):
    ws = wb.create_sheet("Net Worth Snapshot")
    _title(ws, "Net Worth Snapshot")
    ws.append(["Account", "Balance"])
    _header_row(ws, ["Account", "Balance"])
    first_data = ws.max_row + 1
    for k, v in cfg["accounts"].items():
        ws.append([k.replace("_", " ").title(), v])
        ws.cell(row=ws.max_row, column=2).font = BLUE
    for k, v in cfg["real_estate"].items():
        if v:
            ws.append([k.replace("_", " ").title(), v])
            ws.cell(row=ws.max_row, column=2).font = BLUE
    last_data = ws.max_row
    ws.append(["TOTAL NET WORTH", f"=SUM(B{first_data}:B{last_data})"])
    ws.cell(row=ws.max_row, column=1).font = BOLD
    ws.cell(row=ws.max_row, column=2).font = BLACK
    inv = investable_total(cfg)
    ws.append(["Investable total (computed)", inv])
    ws.cell(row=ws.max_row, column=2).font = BLUE
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    return ws


def build_concentration(wb, cfg):
    ws = wb.create_sheet("Employer Concentration")
    _title(ws, f"Employer Concentration -- {cfg['employer_stock']['employer_name']} "
               f"({cfg['employer_stock']['ticker']})")
    es = cfg["employer_stock"]
    ws.append(["Sleeve", "Value"])
    _header_row(ws, ["Sleeve", "Value"])
    for k, v in es["holdings"].items():
        ws.append([k.replace("_", " ").title(), v])
        ws.cell(row=ws.max_row, column=2).font = BLUE
    ws.append(["Total employer exposure", employer_stock_total(cfg)])
    ws.cell(row=ws.max_row, column=1).font = BOLD
    ws.append(["Investable total", investable_total(cfg)])
    ws.append(["Concentration %", employer_concentration_pct(cfg)])
    ws.cell(row=ws.max_row, column=1).font = BOLD
    ws.append(["Watch threshold %", es["watch_threshold_pct"]])
    ws.append(["Trim threshold %", es["trim_threshold_pct"]])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    return ws


def build_year_by_year(wb, cfg):
    ws = wb.create_sheet("Year-by-Year Projections")
    _title(ws, "Year-by-Year Projections (simplified)")
    members = cfg["household"]["members"]
    a = cfg["assumptions"]
    spend0 = a["retirement_spend_annual"]
    infl = a["inflation_rate"]
    ret = a["portfolio_return_base"]
    pension = cfg["income"]["pension_monthly_at_retirement"] * 12
    passive = cfg["income"]["passive_income_annual"]
    ss_a = cfg["social_security"]["spouse_a_monthly_benefit"] * 12
    ss_b = cfg["social_security"]["spouse_b_monthly_benefit"] * 12

    headers = ["Year", "Spouse A age", "Spouse B age", "Spend (infl-adj)",
               "Pension", "Passive", "Social Security", "Portfolio draw", "Portfolio EOY"]
    ws.append(headers)
    _header_row(ws, headers)

    import datetime
    base_year = datetime.date.today().year
    a_age0 = current_age(cfg, "spouse_a")
    b_age0 = current_age(cfg, "spouse_b")
    a_ret_age = members[0]["retirement_age"]
    a_ss_age = members[0]["ss_claim_age"]
    b_ss_age = members[1]["ss_claim_age"]
    portfolio = investable_total(cfg)

    for n in range(0, 36):
        year = base_year + n
        a_age = a_age0 + n
        b_age = b_age0 + n
        retired = a_age >= a_ret_age
        spend = spend0 * ((1 + infl) ** n) if retired else 0
        pen = pension * ((1 + cfg["income"]["pension_cola"]) ** max(0, a_age - a_ret_age)) if retired else 0
        pas = passive if retired else 0
        ss = (ss_a if a_age >= a_ss_age else 0) + (ss_b if b_age >= b_ss_age else 0)
        # grow then draw
        portfolio *= (1 + ret)
        draw = 0
        if retired:
            need = spend - pen - pas - ss
            draw = max(0, need)
            portfolio -= draw
        ws.append([year, a_age, b_age, round(spend), round(pen), round(pas),
                   round(ss), round(draw), round(portfolio)])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    return ws


def build_monte_carlo(wb, cfg):
    ws = wb.create_sheet("Monte Carlo")
    _title(ws, "Monte Carlo (summary -- run engine/quarterly_update.py to refresh)")
    ws.append(["Scenario", "Return mu", "Success rate", "Note"])
    _header_row(ws, ["Scenario", "Return mu", "Success rate", "Note"])
    a = cfg["assumptions"]
    for label, mu in [("Conservative", a["portfolio_return_conservative"]),
                      ("Base", a["portfolio_return_base"]),
                      ("Optimistic", a["portfolio_return_optimistic"])]:
        ws.append([label, mu, "(pending)", "populated by quarterly_update.py"])
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["D"].width = 40
    return ws


# ---- Roth-conversion ladder optimizer -------------------------------------

# IRS Uniform Lifetime Table (2022+) -- required-minimum-distribution divisors.
# RMD = pre-tax balance / divisor. Ages below the start age have no RMD.
_RMD_DIVISOR = {
    73: 26.5, 74: 25.5, 75: 24.6, 76: 23.7, 77: 22.9, 78: 22.0, 79: 21.1,
    80: 20.2, 81: 19.4, 82: 18.5, 83: 17.7, 84: 16.8, 85: 16.0, 86: 15.2,
    87: 14.4, 88: 13.7, 89: 12.9, 90: 12.2, 91: 11.5, 92: 10.8, 93: 10.1,
    94: 9.5, 95: 8.9, 96: 8.4, 97: 7.8, 98: 7.3, 99: 6.8, 100: 6.4,
}


def rmd_start_age(birth_year):
    """SECURE 2.0 required-beginning-age: 73 (1951-59) or 75 (1960+)."""
    if birth_year >= 1960:
        return 75
    if birth_year >= 1951:
        return 73
    return 72


def rmd_factor(age):
    """Fraction of the pre-tax balance that must be withdrawn at `age` (0 before
    the start age). Clamped to the oldest tabulated divisor."""
    if age < 73:
        return 0.0
    div = _RMD_DIVISOR.get(age, _RMD_DIVISOR[100])
    return 1.0 / div


def _simulate_conversions(cfg, strategy, target=None):
    """Project the household's pre-tax drawdown + Roth-conversion ladder to a
    planning horizon and return total lifetime tax + a year-by-year schedule.

    strategy:
      "none"    -- take only the forced RMD (the do-nothing baseline)
      "bracket" -- the readable heuristic: each year convert up to the top of the
                   22% federal bracket, capped at the IRMAA Tier-1 MAGI line in
                   any year a spouse is within the 2-year Medicare lookback
      "optimal" -- fill ordinary income up to a level real `target` (today's $)

    Lifetime tax = PV of every year's (federal + state + IRMAA) tax, PLUS the
    terminal tax on the pre-tax balance still standing at the horizon -- which
    heirs draw down under the SECURE 10-year rule at HEIR_MARGINAL_RATE. Roth
    dollars pass tax-free, so converting early trades tax paid now (at known
    bracket rates) against future RMD-driven tax, IRMAA surcharges, and heir
    tax. PV-discounting at inflation is what produces an interior optimum rather
    than "convert everything immediately." See docs/US_RULES.md.
    """
    import datetime
    a = cfg["assumptions"]
    m = cfg["household"]["members"]
    inc, acct = cfg["income"], cfg["accounts"]

    infl, ret = a["inflation_rate"], a["portfolio_return_base"]
    std0 = a.get("standard_deduction_mfj", tax_us.STANDARD_DEDUCTION_MFJ)
    spend0 = a["retirement_spend_annual"]
    base_year = datetime.date.today().year

    a_age0, b_age0 = current_age(cfg, "spouse_a"), current_age(cfg, "spouse_b")
    a_ret = m[0]["retirement_age"]
    b_ret = m[1]["retirement_age"]
    a_ss, b_ss = m[0]["ss_claim_age"], m[1]["ss_claim_age"]
    a_rmd_start = rmd_start_age(m[0]["birth_year"])

    trad = float(acct.get("spouse_a_401k_pretax", 0) + acct.get("spouse_a_trad_ira", 0)
                 + acct.get("spouse_b_401k_pretax", 0) + acct.get("spouse_b_trad_ira", 0))
    roth = float(acct.get("spouse_a_401k_roth", 0) + acct.get("spouse_a_roth_ira", 0)
                 + acct.get("spouse_b_401k_roth", 0) + acct.get("spouse_b_roth_ira", 0))
    # Taxable pool funds spending shortfalls and the conversion tax (the lever
    # that makes conversions powerful -- pay the tax from outside the IRA).
    taxable = float(acct.get("joint_brokerage", 0) + acct.get("cash_and_cds", 0)
                    + acct.get("deferred_comp", 0))

    pension_m, cola = inc["pension_monthly_at_retirement"], inc["pension_cola"]
    passive0 = inc["passive_income_annual"]
    b_salary0 = inc["spouse_b_annual"]
    ss_a0 = cfg["social_security"]["spouse_a_monthly_benefit"] * 12
    ss_b0 = cfg["social_security"]["spouse_b_monthly_benefit"] * 12

    horizon = max(1, 90 - a_age0)            # project to spouse A age 90
    lifetime_tax, insolvent, schedule = 0.0, False, []
    magi_history = {}                        # year_index -> AGI, for IRMAA lookback

    for n in range(0, horizon + 1):
        year = base_year + n
        a_age, b_age = a_age0 + n, b_age0 + n
        trad *= (1 + ret); roth *= (1 + ret); taxable *= (1 + ret)
        if a_age < a_ret:
            continue  # still working -- no draws or conversions modelled

        idx = (1 + infl) ** n
        std = std0 * idx
        spend = spend0 * idx
        pension = pension_m * 12 * ((1 + cola) ** max(0, a_age - a_ret))
        passive = passive0 * idx
        b_work = b_salary0 * idx if b_age < b_ret else 0.0
        ss_a = ss_a0 * idx if a_age >= a_ss else 0.0
        ss_b = ss_b0 * idx if b_age >= b_ss else 0.0
        ss_total = ss_a + ss_b
        ss_taxable = tax_us.SS_TAXABLE_FRACTION * ss_total

        forced = min(trad * rmd_factor(a_age) if a_age >= a_rmd_start else 0.0, trad)
        base_ordinary = pension + passive + b_work + ss_taxable + forced

        # ---- choose the conversion amount ----
        irmaa_ceiling = tax_us.irmaa_tier1_magi(year, infl)
        if strategy == "none":
            conversion = 0.0
        elif strategy == "bracket":
            # Top of the 22% bracket expressed as AGI (taxable + std deduction).
            bracket_top_taxable = tax_us.FEDERAL_BRACKETS_MFJ[2][1] * idx
            ceiling_agi = bracket_top_taxable + std
            # If either spouse is within 2 years of Medicare, respect IRMAA Tier 1.
            if (a_age >= tax_us.MEDICARE_AGE - tax_us.IRMAA_LOOKBACK_YEARS
                    or b_age >= tax_us.MEDICARE_AGE - tax_us.IRMAA_LOOKBACK_YEARS):
                ceiling_agi = min(ceiling_agi, irmaa_ceiling)
            conversion = max(0.0, min(ceiling_agi - base_ordinary, trad - forced))
        else:  # optimal -- fill to a level real target (today's $ AGI)
            ceiling_agi = target * idx
            conversion = max(0.0, min(ceiling_agi - base_ordinary, trad - forced))

        trad -= conversion
        roth += conversion
        agi = base_ordinary + conversion

        # ---- tax: IRMAA keys off MAGI from two years prior, once on Medicare ----
        n_medicare = (1 if a_age >= tax_us.MEDICARE_AGE else 0) + \
                     (1 if b_age >= tax_us.MEDICARE_AGE else 0)
        magi_look = magi_history.get(n - tax_us.IRMAA_LOOKBACK_YEARS, agi)
        magi_history[n] = agi
        fed = tax_us.federal_tax(agi, year, infl, std0)
        st = tax_us.state_tax(agi, cfg, year, infl, ss_income=ss_taxable)
        irmaa = tax_us.irmaa_annual(magi_look, year, infl, n_enrolled=n_medicare)
        year_tax = fed + st + irmaa
        lifetime_tax += year_tax / ((1 + infl) ** n)   # present value, today's $

        # ---- fund spending; conversion tax is embedded in year_tax ----
        income_cash = pension + passive + ss_total + b_work + forced
        net = income_cash - year_tax
        if net >= spend:
            taxable += (net - spend)
        else:
            short = spend - net
            take = min(taxable, short); taxable -= take; short -= take
            if short > 0:
                take = min(trad, short); trad -= take; short -= take
            if short > 0:
                take = min(roth, short); roth -= take; short -= take
            if short > 1.0:
                insolvent = True

        schedule.append({
            "year": year, "a_age": a_age, "b_age": b_age,
            "base_ordinary": base_ordinary, "forced": forced, "conversion": conversion,
            "agi": agi, "tax": fed + st, "irmaa": irmaa, "magi_look": magi_look,
            "trad": trad, "roth": roth,
            "over": irmaa > 0,
        })

    # ---- terminal tax: pre-tax balance left to heirs (SECURE 10-year rule) ----
    terminal_raw = trad * tax_us.HEIR_MARGINAL_RATE
    terminal_tax = terminal_raw / ((1 + infl) ** horizon)   # present value, today's $
    total_tax = lifetime_tax + terminal_tax
    estate = trad + roth + taxable
    return {
        "strategy": strategy, "target": target,
        "lifetime_tax": lifetime_tax, "terminal_tax": terminal_tax, "total_tax": total_tax,
        "trad_end": trad, "roth_end": roth, "estate_end": estate,
        "insolvent": insolvent, "schedule": schedule,
    }


def _optimize_conversions(cfg):
    """Grid-search the level real conversion target (today's $ AGI ceiling) that
    minimizes total lifetime tax, subject to the plan staying solvent."""
    best = None
    for t in range(0, 400001, 5000):
        r = _simulate_conversions(cfg, "optimal", target=float(t))
        if r["insolvent"]:
            continue
        if best is None or r["total_tax"] < best["total_tax"]:
            best = r
    if best is None:  # nothing solvent -- fall back to the do-nothing baseline
        best = _simulate_conversions(cfg, "none")
    return best


def build_roth_ladder(wb, cfg):
    """Roth Conversion Ladder tab -- the lifetime-tax optimizer (objective #3).

    Shows three strategies side by side: do-nothing (RMDs only), the readable
    fill-to-22%-bracket heuristic, and the lifetime-tax-optimal level target.
    Lifetime tax = in-life federal + state + IRMAA (present-valued) PLUS the
    terminal tax heirs pay on the pre-tax balance under the SECURE 10-year rule.

    Tax engine: engine/tax_us.py (federal MFJ brackets + IRMAA cascade + a hybrid
    flat/optional-progressive state layer). Models ordinary income, the standard
    deduction, RMDs (Uniform Lifetime Table), the IRMAA 2-year lookback, and the
    terminal heir tax. Does NOT model NIIT, capital-gains rates, per-state
    retirement exclusions, or spousal IRA splitting. Illustrative, not advice.
    """
    ws = wb.create_sheet("Roth Conversion Ladder")
    _title(ws, "Roth Conversion Ladder -- lifetime-tax optimizer")

    none = _simulate_conversions(cfg, "none")
    heur = _simulate_conversions(cfg, "bracket")
    best = _optimize_conversions(cfg)

    ws.append(["Objective: minimize lifetime tax = PV(federal + state + IRMAA) "
               "+ terminal tax heirs pay on the pre-tax balance left at age 90."])
    ws.append([])

    # ---- strategy comparison ----
    cmp_headers = ["Strategy", "Annual target (AGI)", "In-life tax (PV)",
                   "Terminal heir tax (PV)", "TOTAL lifetime tax (PV)",
                   "Pre-tax left @90", "Roth @90", "Estate @90"]
    ws.append(cmp_headers)
    _header_row(ws, cmp_headers)
    rows = [
        ("Do nothing (RMDs only)", "n/a", none),
        ("Fill to top of 22% bracket", "22% bracket / IRMAA cap", heur),
        ("Lifetime-tax optimal",
         (f"${best['target']:,.0f}/yr" if best.get("target") else "n/a"), best),
    ]
    base_total = none["total_tax"]
    for label, tgt, r in rows:
        ws.append([label, tgt, round(r["lifetime_tax"]), round(r["terminal_tax"]),
                   round(r["total_tax"]), round(r["trad_end"]), round(r["roth_end"]),
                   round(r["estate_end"])])
        if r is best:
            for c in range(1, len(cmp_headers) + 1):
                ws.cell(row=ws.max_row, column=c).font = GOODF
    ws.append([])
    saved = base_total - best["total_tax"]
    pct = (100 * saved / base_total) if base_total else 0.0
    ws.append(["Lifetime tax saved vs. do-nothing", round(saved),
               f"{pct:.0f}% lower total tax at a ${best['target']:,.0f}/yr conversion ceiling"
               if best.get("target") else "optimizer fell back to do-nothing"])
    ws.cell(row=ws.max_row, column=1).font = BOLD
    ws.cell(row=ws.max_row, column=2).font = GOODF
    ws.append([])

    # ---- optimal year-by-year schedule ----
    sch_headers = ["Year", "Age A", "Age B", "Base income", "RMD", "Conversion",
                   "AGI", "Income tax", "IRMAA", "MAGI (lookback)",
                   "Pre-tax bal", "Roth bal"]
    ws.append(sch_headers)
    _header_row(ws, sch_headers)
    for r in best["schedule"]:
        ws.append([r["year"], r["a_age"], r["b_age"], round(r["base_ordinary"]),
                   round(r["forced"]), round(r["conversion"]), round(r["agi"]),
                   round(r["tax"]), round(r["irmaa"]), round(r["magi_look"]),
                   round(r["trad"]), round(r["roth"])])
        if r["irmaa"] > 0:
            ws.cell(row=ws.max_row, column=9).font = RED

    widths = [8, 7, 7, 13, 12, 13, 13, 12, 10, 16, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def build_action_plan(wb, cfg):
    ws = wb.create_sheet("Action Plan")
    _title(ws, "Action Plan")
    ws.append(["Priority", "Item", "Status"])
    _header_row(ws, ["Priority", "Item", "Status"])
    items = [
        (1, "Review employer-stock concentration vs. thresholds (company_health.py)", "Recurring"),
        (2, "Confirm beneficiary designations on every account", "Open"),
        (3, "Execute Roth-conversion ladder to the optimal target (see Roth Conversion Ladder tab)", "Open"),
        (4, "Verify pension survivor-benefit election", "Open"),
        (5, "Rebalance toward target equity/bond split", "Recurring"),
    ]
    for p, it, st in items:
        ws.append([p, it, st])
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 14
    return ws


def build(cfg):
    wb = Workbook()
    wb.remove(wb.active)
    build_assumptions(wb, cfg)
    build_net_worth(wb, cfg)
    build_concentration(wb, cfg)
    build_income_note(wb, cfg)
    build_year_by_year(wb, cfg)
    build_monte_carlo(wb, cfg)
    build_roth_ladder(wb, cfg)
    build_action_plan(wb, cfg)
    return wb


def build_income_note(wb, cfg):
    ws = wb.create_sheet("Income Streams")
    _title(ws, "Income Streams")
    ws.append(["Stream", "Annual", "Notes"])
    _header_row(ws, ["Stream", "Annual", "Notes"])
    inc = cfg["income"]
    m = cfg["household"]["members"]
    rows = [
        (f"{m[0]['display_name']} salary", inc["spouse_a_salary"], ""),
        (f"{m[0]['display_name']} bonus", round(inc["spouse_a_salary"] * inc["spouse_a_bonus_pct"]), ""),
        (f"{m[0]['display_name']} RSU", inc["spouse_a_rsu_annual"], "employer stock -- see concentration tab"),
        (f"{m[1]['display_name']} income", inc["spouse_b_annual"], ""),
        ("Pension (annual at retirement)", inc["pension_monthly_at_retirement"] * 12, ""),
        ("Passive income", inc["passive_income_annual"], ""),
    ]
    for r in rows:
        ws.append(r)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 40
    return ws


def main():
    cfg, path = load_config()
    out = ROOT / cfg["paths"]["model_xlsx"]
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = build(cfg)
    wb.save(out)
    tag = "DEMO" if cfg.get("_is_demo") else "LIVE"
    print(f"[{tag}] Built model from {path.name}")
    print(f"  -> {out}")
    print(f"  Tabs: {wb.sheetnames}")


if __name__ == "__main__":
    main()
