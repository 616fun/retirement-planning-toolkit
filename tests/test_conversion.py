"""Unit tests for the Roth-conversion ladder optimizer and RMD logic."""
import copy
import pathlib

import openpyxl
import pytest
import config_loader as cl
import build_model as bm

ROOT = pathlib.Path(__file__).resolve().parent.parent


def _cfg():
    cfg, _ = cl.load_config(str(ROOT / "config" / "examples" / "rivera_config.json"))
    return cfg


# ---- RMD table -----------------------------------------------------------
def test_rmd_start_age_by_birth_year():
    assert bm.rmd_start_age(1976) == 75      # born 1960+
    assert bm.rmd_start_age(1955) == 73      # born 1951-59
    assert bm.rmd_start_age(1945) == 72


def test_rmd_factor_zero_before_start_then_climbs():
    assert bm.rmd_factor(70) == 0.0
    assert bm.rmd_factor(72) == 0.0          # table begins at 73
    assert bm.rmd_factor(75) == pytest.approx(1 / 24.6, abs=1e-6)
    assert bm.rmd_factor(90) == pytest.approx(1 / 12.2, abs=1e-6)
    assert bm.rmd_factor(130) == pytest.approx(1 / 6.4, abs=1e-6)   # clamps to oldest


# ---- simulation invariants ----------------------------------------------
def test_simulation_invariants():
    r = bm._simulate_conversions(_cfg(), "optimal", target=150000.0)
    assert r["lifetime_tax"] > 0
    assert all(row["trad"] >= -1e-6 for row in r["schedule"])     # never goes negative
    assert all(row["conversion"] >= -1e-6 for row in r["schedule"])
    assert r["terminal_tax"] >= 0


def test_conversions_move_pretax_into_roth():
    none = bm._simulate_conversions(_cfg(), "none")
    opt = bm._simulate_conversions(_cfg(), "optimal", target=150000.0)
    # Converting shrinks the pre-tax balance at the horizon and grows Roth.
    assert opt["trad_end"] < none["trad_end"]
    assert opt["roth_end"] > none["roth_end"]


def _modest_cfg():
    # A smaller-balance, lower-spend household where the ACA cliff binds.
    cfg = _cfg()
    cfg["accounts"].update(spouse_a_401k_pretax=180000, spouse_a_trad_ira=60000,
                           spouse_b_401k_pretax=120000, spouse_b_trad_ira=40000,
                           joint_brokerage=250000, cash_and_cds=60000)
    cfg["assumptions"]["retirement_spend_annual"] = 85000
    cfg["healthcare"]["aca_benchmark_premium_annual"] = 20000
    return cfg


# ---- optimizer -----------------------------------------------------------
def test_optimal_beats_do_nothing():
    best = bm._optimize_conversions(_cfg())
    none = bm._simulate_conversions(_cfg(), "none")
    assert best["net_cost"] <= none["net_cost"]
    assert best["target"] is not None
    assert not best["insolvent"]


def test_heuristic_is_competitive_with_optimal():
    # The fill-to-bracket heuristic should land near (never below) the optimum.
    heur = bm._simulate_conversions(_cfg(), "bracket")
    best = bm._optimize_conversions(_cfg())
    assert best["net_cost"] <= heur["net_cost"] + 1.0        # optimal is the floor
    assert heur["net_cost"] <= 1.5 * best["net_cost"]        # heuristic within 50%


def test_optimizer_is_deterministic():
    a = bm._optimize_conversions(_cfg())
    b = bm._optimize_conversions(_cfg())
    assert a["target"] == b["target"]
    assert a["net_cost"] == pytest.approx(b["net_cost"])


# ---- ACA awareness (the Phase 1 correctness fix) -------------------------
def test_net_cost_subtracts_aca_subsidy():
    r = bm._simulate_conversions(_modest_cfg(), "none")   # low-MAGI baseline keeps ACA
    assert r["lifetime_aca_subsidy"] > 0
    assert r["net_cost"] == pytest.approx(r["total_tax"] - r["lifetime_aca_subsidy"])


def test_aca_awareness_changes_the_optimum():
    on = _modest_cfg()
    off = copy.deepcopy(on)
    off["healthcare"]["aca_benchmark_premium_annual"] = 0   # ACA-blind
    b_on = bm._optimize_conversions(on)
    b_off = bm._optimize_conversions(off)
    assert b_on["lifetime_aca_subsidy"] > 0          # the optimum preserves real subsidy
    assert b_on["target"] != b_off["target"]         # which the ACA-blind run ignores


def test_aca_cliff_costs_subsidy_in_simulation():
    # A moderate level conversion parks MAGI just over the 400% FPL cliff every
    # pre-65 year, forfeiting the subsidy the low-MAGI do-nothing plan keeps.
    # (Bunching conversions into one year would preserve more -- a time-varying
    # schedule the level-target optimizer can only approximate; see Phase 2.)
    cfg = _modest_cfg()
    none = bm._simulate_conversions(cfg, "none")
    midconv = bm._simulate_conversions(cfg, "optimal", target=70000.0)
    assert none["lifetime_aca_subsidy"] > midconv["lifetime_aca_subsidy"]


def test_no_income_tax_state_costs_less_than_taxed_state():
    base = _cfg()
    notax = copy.deepcopy(base)
    notax["household"]["state_income_tax_rate"] = 0.0
    notax["household"]["local_income_tax_rate"] = 0.0
    taxed = copy.deepcopy(base)
    taxed["household"]["state_income_tax_rate"] = 0.05
    taxed["household"]["local_income_tax_rate"] = 0.0
    assert (bm._simulate_conversions(notax, "optimal", target=150000.0)["lifetime_tax"]
            < bm._simulate_conversions(taxed, "optimal", target=150000.0)["lifetime_tax"])


# ---- workbook tab --------------------------------------------------------
def test_ladder_tab_present_and_populated(tmp_path):
    wb = bm.build(_cfg())
    assert "Roth Conversion Ladder" in wb.sheetnames
    out = tmp_path / "plan.xlsx"
    wb.save(out)
    ws = openpyxl.load_workbook(out)["Roth Conversion Ladder"]
    cells = [c.value for row in ws.iter_rows() for c in row]
    assert any(isinstance(v, str) and "Net-cost optimal" in v for v in cells)
    assert any(isinstance(v, str) and "Do nothing" in v for v in cells)
