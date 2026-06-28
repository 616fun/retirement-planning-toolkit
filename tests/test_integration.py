"""Integration tests: the demo builds the full workbook, renders the dashboard,
and runs the Monte Carlo pipeline without error."""
import pathlib

import openpyxl
import pytest
import config_loader as cl
import build_model as bm
import refresh_dashboard as rd
import quarterly_update as qu

ROOT = pathlib.Path(__file__).resolve().parent.parent
EXPECTED_SHEETS = {
    "Assumptions", "Net Worth Snapshot", "Employer Concentration", "Income Streams",
    "Year-by-Year Projections", "Monte Carlo", "Roth Conversion Ladder", "Action Plan",
}
DEMO = "rivera_config.json"


def _cfg(name=DEMO):
    cfg, _ = cl.load_config(str(ROOT / "config" / "examples" / name))
    return cfg


# ---- workbook ------------------------------------------------------------
def test_build_produces_all_tabs():
    wb = bm.build(_cfg())
    assert set(wb.sheetnames) == EXPECTED_SHEETS


def test_built_workbook_round_trips(tmp_path):
    wb = bm.build(_cfg())
    out = tmp_path / "plan.xlsx"
    wb.save(out)
    reloaded = openpyxl.load_workbook(out)
    assert EXPECTED_SHEETS.issubset(set(reloaded.sheetnames))
    # the net-worth tab should carry a TOTAL NET WORTH line
    ws = reloaded["Net Worth Snapshot"]
    cells = [c.value for row in ws.iter_rows() for c in row]
    assert any(isinstance(v, str) and "TOTAL NET WORTH" in v for v in cells)


# ---- dashboard -----------------------------------------------------------
def test_dashboard_renders_with_panels():
    cfg = _cfg()
    doc = rd.render(cfg, qu.monte_carlo(cfg, n_sims=300))
    assert "<!DOCTYPE html>" in doc
    assert 'lang="en"' in doc
    assert cfg["household"]["name"] in doc
    assert 'class="kpis"' in doc            # KPI tiles present
    assert doc.count('class="panel"') >= 2  # concentration + Monte Carlo panels


def test_dashboard_renders_without_monte_carlo():
    # Should degrade gracefully when no MC summary is supplied.
    html = rd.render(_cfg(), mc=None)
    assert "<!DOCTYPE html>" in html
    assert "Run quarterly_update.py" in html  # MC placeholder copy


# ---- Monte Carlo ---------------------------------------------------------
def test_monte_carlo_deterministic_and_in_range():
    cfg = _cfg()
    m1 = qu.monte_carlo(cfg, n_sims=500)
    m2 = qu.monte_carlo(cfg, n_sims=500)
    for k in ("conservative", "base", "optimistic"):
        assert 0.0 <= m1[k]["success_rate"] <= 100.0
        assert m1[k]["success_rate"] == m2[k]["success_rate"]   # fixed seed -> reproducible
    assert m1["conservative"]["success_rate"] <= m1["optimistic"]["success_rate"]


def test_mc_stamps_into_workbook(tmp_path):
    cfg = _cfg()
    out = tmp_path / "plan.xlsx"
    bm.build(cfg).save(out)
    qu.stamp_mc(str(out), qu.monte_carlo(cfg, n_sims=200))
    ws = openpyxl.load_workbook(out)["Monte Carlo"]
    stamped = {r[0].value: r[2].value for r in ws.iter_rows(min_col=1, max_col=4)}
    assert stamped.get("Base", "").endswith("%")


# ---- quarterly overlay ---------------------------------------------------
def test_quarterly_apply_input_overlays_balances():
    cfg = _cfg()
    before = cfg["accounts"]["spouse_a_401k_pretax"]
    cfg2, msg = qu.apply_input(cfg, None)   # no input file -> unchanged
    assert cfg2["accounts"]["spouse_a_401k_pretax"] == before
